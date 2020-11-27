# -*- coding:utf-8 -*-
import openpyxl
import sys
from datetime import datetime
import xlrd
import csv
import os
from productoGPG import Producto_GPG
import strings      # Cadenas de caracteres constantes
import copy

# Tuplas de posibilidades
tipos = ('variable','variation','simple')

# Tupla de campos posibles en archivo destino
tupla_campos_destino = ('ID','Tipo','SKU','Nombre','Publicado','¿Está destacado?','Visibilidad en el catálogo','Descripción corta','Descripción','Día en que empieza el precio rebajado','Día en que termina el precio rebajado','Estado del impuesto','Clase de impuesto','¿En inventario?','Inventario','Cantidad de bajo inventario','¿Permitir reservas de productos agotados?','¿Vendido individualmente?','Peso (kg)','Longitud (cm)','Anchura (cm)','Altura (cm)','¿Permitir valoraciones de clientes?','Nota de compra','Precio rebajado','Precio normal','Categorías','Etiquetas','Clase de envío','Imágenes','Límite de descargas','Días de caducidad de la descarga','Superior','Productos agrupados','Ventas dirigidas','Ventas cruzadas','URL externa','Texto del botón','Posición','Fixed Tiered Prices','Nombre del atributo 1','Valor(es) del atributo 1','Atributo visible 1','Atributo global 1','Meta: wcz_pps_price_prefix','Meta: site-sidebar-layout','Meta: site-content-layout','Meta: theme-transparent-header-meta','Meta: precio_menudeo','Meta: _precio_menudeo','Meta: precio_mayoreo','Meta: _precio_mayoreo','Nombre del atributo 2','Valor(es) del atributo 2','Atributo visible 2','Atributo global 2','Meta: _wp_page_template','Nombre del atributo 3','Valor(es) del atributo 3','Atributo visible 3','Atributo global 3')

# Tupla de campos posibles en archivo origen
tupla_campos_origen = ('ID','SKU','Nombre','Descripción','Descripción corta','Precio normal ','Categorías','Familia SKU (productos de la misma línea)','Mayoreo a partir de (1 indica no mayoreo)','Precio de mayoreo (si está considerado)','¿En existencia? (1- si, 0-no)','Marca','Atributo','Cantidad','Unidad','Archivo de imagen (URL)')

# Diccionario de columnas origen
campos_origen = {}

# Diccionario de columnas destino
campos_destino = {}

#Lista con nombres de las columnas origen
nombres_columnas_origen = []

#Lista con nombres de las columnas destino
nombres_columnas_destino = []

#Listas para objetos de tipo Producto
lista_productos = []   
lista_productos_padres = []
lista_productos_hijos = []
lista_productos_simples = []

lista_sku_general = []              # Lista general de SKUs
lista_sku_familias = []             # Lista de familias de SKU

lista_productos_final = []

##
def get_column_names(worksheet):
    names=[]
    i=0
    for i in range(1,get_total_columns_sheet(worksheet)+1):
        cell = worksheet.cell(row=1, column=i)
        names.append(cell.value)
    return names

def get_column_number(worksheet,column_id):
    i = 1   #iterador
    while(worksheet.cell(row=1,column=i).value != column_id):
        i += 1
    return i  

def get_total_rows_sheet(worksheet):
    return worksheet.max_row

def get_total_columns_sheet(worksheet):
    return worksheet.max_column

def get_total_registers(worksheet):
    return get_total_rows_sheet(worksheet)-1

def get_register_by_id(worksheet,id):
    reg = []
    i=0
    j=0
    for i in range(0,get_total_rows_sheet(worksheet)+1):
        if( i == id):
            break
    i+=2
        
    for j in range(1,get_total_columns_sheet(worksheet)+1):
        reg.append(worksheet.cell(row=i,column=j).value)
    return reg

# Función de obtención de la lista de todos los productos en el catálogo actual
def get_productos(worksheet):
    l_productos = []
    l_atributos = []
    print('Obteniendo productos')
    global nombres_columnas_origen
    try:
        for i in range(2,get_total_rows_sheet(worksheet)+1):
            #print('i: '+str(i))
            columnas_origen_len = len(nombres_columnas_origen)
            #print('len columnas: '+str(columnas_len))
            for j in range(columnas_origen_len):
                #print('j: '+str(j))
                nombre_columna = nombres_columnas_origen[j]
                atributo = str(worksheet.cell(row=i, column=get_key(nombre_columna,campos_origen)).value)
                #print(nombre_columna+': '+atributo)
                l_atributos.append(atributo)
            #print_lista(l_atributos)
            l_atributos.reverse()
            #print_lista(atributos)
            #print(len(atributos))
            id = l_atributos.pop()
            sku = l_atributos.pop()
            nom = l_atributos.pop()
            desc = l_atributos.pop()
            desc_c = l_atributos.pop()
            pu = l_atributos.pop()
            cat = l_atributos.pop()
            fam = l_atributos.pop()
            may = int(l_atributos.pop())
            pm = float(l_atributos.pop())
            exis = int(l_atributos.pop())
            mar = l_atributos.pop()
            atr = l_atributos.pop()
            cnt = int(l_atributos.pop())
            uni = l_atributos.pop()
            img = l_atributos.pop()
            producto = Producto_GPG(id,sku,nom,desc,desc_c,pu,cat,fam,may,pm,exis,mar,atr,cnt,uni,img)

            l_atributos.clear()

            # producto.print_prod()  #Impresión de atributos de cada producto (debug)
            l_productos.append(producto)
        return l_productos
    except Exception as ex:
        print('Excepción: '+str(ex))
        return []


# Función para imprimir los elementos de una lista predefinidoa
def print_lista(lista):
    if len(lista) == 0:
        print('Lista vacía')
    else:
        for elemento in lista:
            print(elemento)


# Función para imprimir los productos de una lista predefinidoa
def print_productos(lista):
    if len(lista) == 0:
        print('Lista vacía')
    else:
        i = 1
        for producto in lista:
            print(str(i),end=' ')
            producto.print_prod()
            i += 1

# Obtención de lista de productos simples (no son padres ni hijos)
def get_productos_simples(lista_productos):
    lista = [producto for producto in lista_productos if producto.tipo=='simple']
    return lista

# Obtención de lista de productos padres
def get_productos_padres(lista_productos):
    lista = [producto for producto in lista_productos if producto.tipo=='variable']
    return lista

# Obtención de lista de productos hijos
def get_productos_hijos(lista_productos):
    lista = [producto for producto in lista_productos if producto.tipo=='variation']
    return lista

# Función para devolver la clave para cualquier valor  en un diccionario
def get_key(val,dicc): 
    for key, value in dicc.items(): 
        if val == value: 
            return key 
    return 'La clave no existe'

# función que devuelve lista completa de productos de la familia SKU indicada
def get_familia(sku_familia):
    familia = [producto for producto in lista_productos if producto.familia==sku_familia]            
    return familia

# Función para llenar la primera fila con una lista o tupla de campos definida
def llena_primera_fila(worksheet,data):
    i = 1
    for campo in data:
        celda = worksheet.cell(row = 1, column=i)
        celda.value = campo
        i += 1

# Función para conversión de archivo .xlsx a .csv
def csv_from_excel(excel_file,csv_file):
    workbook = xlrd.open_workbook(excel_file, encoding_override='utf-8')
    sheet = workbook.sheet_by_name('Sheet')
    csv_file = open(csv_file, 'w')
    wr = csv.writer(csv_file, quoting=csv.QUOTE_ALL, lineterminator='\n')
    for rownum in range(sheet.nrows):
        wr.writerow(sheet.row_values(rownum))
    csv_file.close()

#
def ocurrencias_valor_lista(valor,lista):
    ocurrencias = 0
    for elemento in lista:
        if elemento == valor:
            ocurrencias += 1
    return ocurrencias

# Función principal
#def main():
sys.stdout.reconfigure(encoding='utf-8')

# Obtención de fecha y hora actual para generar archivos distintos cada vez que se ejecuta el programa
fecha_hora = datetime.now()
fecha_hora_cad = fecha_hora.strftime("%Y%m%d_%H%M%S")

nombre_archivo_origen = sys.argv[1]       # El argumento a pasar al programa es la ruta del archivo origen
ruta_archivo_origen = nombre_archivo_origen
directorio_archivos_destino = 'carga masiva'

if(os.path.isdir(directorio_archivos_destino)):
    print('El directorio de archivos de carga masiva ya existe')
else:
    try:
        os.mkdir(directorio_archivos_destino)
        print('Directorio de archivos de carga masiva creado')
    except Exception as e:
        print('Error al crear directorio: '+str(e))

ruta_archivo_destino = directorio_archivos_destino+'/carga_masiva_'+fecha_hora_cad+'.xlsx'
ruta_archivo_destino_csv = directorio_archivos_destino+'/carga_masiva_'+fecha_hora_cad+'.csv'

print('Archivo origen: '+ruta_archivo_origen)
print('Archivo destino .xlsx: '+ruta_archivo_destino)
print('Archivo destino .csv: '+ruta_archivo_destino_csv)

print('Abriendo hojas de cálculo')

# for campo in campos_origen:
#     print(campo)


try:
    workbook_origen = openpyxl.load_workbook(ruta_archivo_origen)
    workbook_destino = openpyxl.Workbook()
    
    print('Hojas de cálculo abiertas')
    sheet_origen = workbook_origen.active                                     # Objetos de tipo hoja dentro de archivo xlsx
    sheet_destino = workbook_destino.active
    llena_primera_fila(sheet_destino,tupla_campos_destino)

    print('Cantidad de registros: '+str(get_total_registers(sheet_origen)))
    
    nombres_columnas_origen = get_column_names(sheet_origen)                      # Obtención de nombres de las columnas
    #print('Columnas origen: ')
    #print_lista(nombres_columnas_origen)

    nombres_columnas_destino = get_column_names(sheet_destino)
    #print('Columnas destino: ')
    #print_lista(nombres_columnas_destino)

    campos_origen = dict(zip(range(1,len(nombres_columnas_origen)+1),nombres_columnas_origen))   # Diccionario de todos los nombres de campos de origen posibles
    campos_origen_len = len(campos_origen.values())
    #print('Campos origen: '+str(campos_origen_len))
    #print(campos_origen.keys())
    #print(campos_origen.values())

    campos_destino = dict(zip(range(1,len(nombres_columnas_destino)+1),nombres_columnas_destino))   # Diccionario de todos los nombres de campos de destino posibles
    campos_destino_len = len(campos_destino.values())
    #print('Campos destino: '+str(campos_destino_len))
    #print(campos_destino.keys())
    #print(campos_destino.values())



    lista_productos = get_productos(sheet_origen)                          # Obtención de lista general de productos 
    print('Cantidad total de productos: '+str(len(lista_productos)))  

    #print_productos(lista_productos)                                   # Imprimir todos los productos (opcional para fines de depuración)

    lista_sku_general = [item.familia for item in lista_productos]  # Obtención de todos los SKU's
    lista_sku_general_len = len(lista_sku_general)
    #print('Cantidad total SKU: '+str(lista_sku_general_len))

    lista_sku_familias = list(set(lista_sku_general))                   # Obtención de familias de SKU's
    lista_sku_familias_len = len(lista_sku_familias)
    print('Cantidad familias SKU: '+str(lista_sku_familias_len))


    # Recuento de familias y generación de listas
    for sku_familia in lista_sku_familias:
        familia_actual = get_familia(sku_familia)
        familia_actual_len = len(familia_actual)


        #Determinación de tipo de producto
        for producto in familia_actual:
            if(familia_actual_len==1):                                  # Simple
                producto.tipo = 'simple'
                #producto.sku = str(sku_familia)
                producto.permitir_valoraciones_clientes = '1'
                producto.imagenes = producto.url_img
            else:                
                if producto.descripcion_corta == 'None':                # Variation
                    producto.tipo = 'variation'
                    producto.clase_impuesto = 'parent'
                    producto.permitir_valoraciones_clientes = '0'
                    producto.imagenes = producto.url_img
                    producto.superior = producto.familia

                    lista_sku = []
                    for prod in familia_actual:
                        lista_sku.append(prod.sku)
                    if(ocurrencias_valor_lista(producto.sku,lista_sku) == 1):
                        pass
                    else:
                        prefijo = ('2-' if 'CAJA' in producto.unidad else ('3-' if 'EMPAQUE' in producto.unidad or 'EMP' in producto.unidad else ('' if 'PIEZA' in producto.unidad or 'Pieza' in producto.unidad else '')))
                        producto.sku = prefijo+producto.sku
                else:
                    producto.tipo = 'variable'                          # Variable
                    #producto.sku = str(sku_familia)
                    producto.permitir_valoraciones_clientes = '1'
                    producto_copia = copy.deepcopy(producto)   # Copia de producto base mediante deepcopy

                    producto_copia.tipo = 'variation'
                    producto_copia.sku = ''
                    producto_copia.descripcion_corta = 'None'
                    producto_copia.permitir_valoraciones_clientes = '0'
                    producto_copia.precio_normal = ''
                    producto_copia.imagenes = producto.url_img
                    producto_copia.superior = producto_copia.familia
                    producto_copia.categorias = ''
                    familia_actual.append(producto_copia)

        #Ciclo auxiliar para sustituir campos 'None' por vacíos
        for producto in familia_actual:
            if producto.descripcion_corta == 'None':
                    producto.descripcion_corta = ''
            if producto.categorias == 'None':
                    producto.categorias = ''
            if producto.sku == 'None':
                producto.sku = ''
            # producto.print_prod()

        ### Obtención de imágenes
        for producto in familia_actual:
            if(producto.tipo == 'variable'):
                fam = get_familia(producto.familia)
                for prod in fam:
                    producto.imagenes += prod.url_img+', '
            
        # Obtención de posiciones
        contador_posicion = 0
        for producto in familia_actual:
            if producto.tipo == 'simple':
                producto.posicion = 0
                contador_posicion = 0
            elif producto.tipo == 'variable':
                producto.posicion = 0
                contador_posicion += 1
            else:   # Variation
                producto.posicion = contador_posicion
                contador_posicion += 1

        #Obtención de atributo 1
        for producto in familia_actual:
            if producto.tipo == 'simple' or producto.tipo == 'variable':
                producto.nombre_atributo_1 = 'MARCA'
                producto.valor_atributo_1 = producto.marca
                producto.atributo_visible_1 = '1'
                producto.atributo_global_1 = '1'
            elif producto.tipo == 'variation' :
                producto.valor_atributo_1 = producto.unidad
                producto.atributo_visible_1 = ''
                producto.meta_site_sidebar_layout = ''
                producto.meta_site_content_layout = ''
                producto.meta_theme_transparent_header_meta = ''

        # Obtención de tamaño de familia (después de haber creado copia del padre si fue el caso)
        familia_size = len(familia_actual)
        #print('Tamaño de familia: '+str(familia_size))

        # Obtención de atributo 2
        for producto in familia_actual:
            producto.nombre_atributo_2 = 'PRESENTACION' if producto.tipo == 'simple' or producto.tipo == 'variable' else 'ATRIBUTO' 
            producto.atributo_visible_2 = '1' if producto.tipo == 'simple' or producto.tipo == 'variable' else ''
            if producto.tipo == 'simple':
                producto.valor_atributo_2 = producto.unidad
            if producto.tipo == 'variable':
                fam = get_familia(producto.familia)
                for prod in fam:
                    producto.valor_atributo_2 += prod.atributo +', '
            if producto.tipo == 'variation':
                producto.valor_atributo_2 = producto.atributo if familia_actual_len > 3 else ''
    
        # Obtención de atributo 3
        for producto in familia_actual:
            producto.nombre_atributo_3 = ('ATRIBUTO' if familia_actual_len > 3 else '') if producto.tipo == 'variable' else ''
            if producto.tipo == 'variable':
                producto.atributo_visible_3 = '1' if familia_actual_len > 3 else '0'
                producto.atributo_global_3 = '0'
            elif producto.tipo == 'simple':
                producto.atributo_visible_3 = '0'
                producto.atributo_global_3 = '0'




        # Obtención de meta_page_template
        for producto in familia_actual:
            producto.meta_wp_page_template = 'default' if producto.tipo == 'variable' else ''

        #Obtención de rangeo
        for producto in familia_actual:
            if(producto.mayoreo_a_partir != 1):
                producto.meta_precio_mayoreo = 'A partir de '+str(producto.mayoreo_a_partir)+' unidades'
                producto.meta_precio_menudeo = 'de 1 a '+str(producto.mayoreo_a_partir-1)+' unidades'
                producto.fixed_tiered_prices = str(producto.mayoreo_a_partir)+':'+str(producto.precio_mayoreo)

        # Llenado de lista final de productos
        for producto in familia_actual:
            lista_productos_final.append(producto)
    
    #print_productos(lista_productos_final)

    # Llenado de archivo de carga masiva
    contador_filas = 2      # Conteo de filas inicia en la segunda
    for producto in lista_productos_final:
        celda = sheet_destino.cell(row = contador_filas, column=get_key('ID',campos_destino))
        celda.value = producto.id
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Tipo',campos_destino))
        celda.value = producto.tipo
        celda = sheet_destino.cell(row = contador_filas, column=get_key('SKU',campos_destino))
        celda.value = producto.sku
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Nombre',campos_destino))
        celda.value = producto.nombre
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Publicado',campos_destino))
        celda.value = producto.publicado
        celda = sheet_destino.cell(row = contador_filas, column=get_key('¿Está destacado?',campos_destino))
        celda.value = producto.destacado
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Visibilidad en el catálogo',campos_destino))
        celda.value = producto.visibilidad_catalogo
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Descripción corta',campos_destino))
        celda.value = producto.descripcion_corta
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Descripción',campos_destino))
        celda.value = producto.descripcion
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Día en que empieza el precio rebajado',campos_destino))
        celda.value = producto.dia_inicio_rebaja
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Día en que termina el precio rebajado',campos_destino))
        celda.value = producto.dia_fin_rebaja
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Estado del impuesto',campos_destino))
        celda.value = producto.estado_impuesto
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Clase de impuesto',campos_destino))
        celda.value = producto.clase_impuesto
        celda = sheet_destino.cell(row = contador_filas, column=get_key('¿En inventario?',campos_destino))
        celda.value = producto.en_inventario
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Inventario',campos_destino))
        celda.value = producto.inventario
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Cantidad de bajo inventario',campos_destino))
        celda.value = producto.cantidad_bajo_inventario
        celda = sheet_destino.cell(row = contador_filas, column=get_key('¿Permitir reservas de productos agotados?',campos_destino))
        celda.value = producto.permitir_reserva_producto_agotado
        celda = sheet_destino.cell(row = contador_filas, column=get_key('¿Vendido individualmente?',campos_destino))
        celda.value = producto.vendido_individualmente
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Peso (kg)',campos_destino))
        celda.value = producto.peso_kg
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Longitud (cm)',campos_destino))
        celda.value = producto.longitud_cm
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Anchura (cm)',campos_destino))
        celda.value = producto.anchura_cm
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Altura (cm)',campos_destino))
        celda.value = producto.altura_cm
        celda = sheet_destino.cell(row = contador_filas, column=get_key('¿Permitir valoraciones de clientes?',campos_destino))
        celda.value = producto.permitir_valoraciones_clientes
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Nota de compra',campos_destino))
        celda.value = producto.nota_de_compra
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Precio rebajado',campos_destino))
        celda.value = producto.precio_rebajado
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Precio normal',campos_destino))
        celda.value = producto.precio_normal
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Categorías',campos_destino))
        celda.value = producto.categorias
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Etiquetas',campos_destino))
        celda.value = producto.etiquetas
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Clase de envío',campos_destino))
        celda.value = producto.clase_envio
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Imágenes',campos_destino))
        celda.value = producto.imagenes
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Límite de descargas',campos_destino))
        celda.value = producto.limite_descargas
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Días de caducidad de la descarga',campos_destino))
        celda.value = producto.dias_caducidad
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Superior',campos_destino))
        celda.value = producto.superior
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Productos agrupados',campos_destino))
        celda.value = producto.productos_agrupados
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Ventas dirigidas',campos_destino))
        celda.value = producto.ventas_dirigidas
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Ventas cruzadas',campos_destino))
        celda.value = producto.ventas_cruzadas
        celda = sheet_destino.cell(row = contador_filas, column=get_key('URL externa',campos_destino))
        celda.value = producto.url_externa
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Texto del botón',campos_destino))
        celda.value = producto.texto_boton
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Posición',campos_destino))
        celda.value = producto.posicion
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Fixed Tiered Prices',campos_destino))
        celda.value = producto.fixed_tiered_prices
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Nombre del atributo 1',campos_destino))
        celda.value = producto.nombre_atributo_1
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Valor(es) del atributo 1',campos_destino))
        celda.value = producto.valor_atributo_1
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Atributo visible 1',campos_destino))
        celda.value = producto.atributo_visible_1
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Atributo global 1',campos_destino))
        celda.value = producto.atributo_global_1
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Meta: wcz_pps_price_prefix',campos_destino))
        celda.value = producto.meta_wcz_pps_price_prefix
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Meta: site-sidebar-layout',campos_destino))
        celda.value = producto.meta_site_sidebar_layout
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Meta: site-content-layout',campos_destino))
        celda.value = producto.meta_site_content_layout
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Meta: theme-transparent-header-meta',campos_destino))
        celda.value = producto.meta_theme_transparent_header_meta
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Meta: precio_menudeo',campos_destino))
        celda.value = producto.meta_precio_menudeo
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Meta: _precio_menudeo',campos_destino))
        celda.value = producto.meta__precio_menudeo
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Meta: precio_mayoreo',campos_destino))
        celda.value = producto.meta_precio_mayoreo
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Meta: _precio_mayoreo',campos_destino))
        celda.value = producto.meta__precio_mayoreo
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Nombre del atributo 2',campos_destino))
        celda.value = producto.nombre_atributo_2
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Valor(es) del atributo 2',campos_destino))
        celda.value = producto.valor_atributo_2
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Atributo visible 2',campos_destino))
        celda.value = producto.atributo_visible_2
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Atributo global 2',campos_destino))
        celda.value = producto.atributo_global_2
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Meta: _wp_page_template',campos_destino))
        celda.value = producto.meta_wp_page_template
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Nombre del atributo 3',campos_destino))
        celda.value = producto.nombre_atributo_3
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Valor(es) del atributo 3',campos_destino))
        celda.value = producto.valor_atributo_3
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Atributo visible 3',campos_destino))
        celda.value = producto.atributo_visible_3
        celda = sheet_destino.cell(row = contador_filas, column=get_key('Atributo global 3',campos_destino))
        celda.value = producto.atributo_global_3
        ### Delimitador
        contador_filas += 1     # Incremento de fila

        #DEBUG
        # for producto in familia_actual:
        #     print(producto.tipo +'-'+producto.unidad+' '+producto.sku) 

    # Funciones de finalización (guardado de archivos finales)
    workbook_destino.save(ruta_archivo_destino)                             # Guarda hoja de cálculo destino
    csv_from_excel(ruta_archivo_destino,ruta_archivo_destino_csv)           # Obtención de .csv a partir de .xlsx
    if os.path.exists(ruta_archivo_destino):
        os.remove(ruta_archivo_destino)
    else:
        print('Error, el archivo no existe')


# contador_posicion = 0
# for producto in lista_productos:
#     #Copia de padre?
#     if(producto.sku == 'None'):
#         producto.tipo = 'variation'
#         producto.sku_new = ''   # Sin SKU
#         producto.superior_new = producto.superior
#         producto.posicion = contador_posicion
#         contador_posicion = 0
#     else:        
#         familia_actual = get_familia(producto.superior)
#         if len(familia_actual)==1:  # verifica si la familia es de un solo elemento (producto de tipo 'simple')
#             producto.tipo = 'simple'
#             producto.sku_new = str(producto.sku)        # Sin cambios en SKU
#             producto.superior_new = 'xxx'
#             producto.posicion = 0
#             contador_posicion = 0
#         else:
#             if producto in familia_actual:
#                 if(producto.sku == producto.superior):
#                     if(producto.clasificacion == '1-PZA'):
#                         producto.sku_new = str(producto.sku)        # Sin cambios en SKU
#                         producto.tipo = 'variable'
#                         producto.superior_new = 'xxx'
#                         producto.posicion = 0
#                         contador_posicion = 1
#                     elif(producto.clasificacion == '2-CAJILLA'):
#                         producto.sku_new = '2-'+str(producto.sku)
#                         producto.tipo = 'variation'
#                         producto.superior_new = producto.superior
#                         producto.posicion = contador_posicion
#                         contador_posicion += 1
#                     elif(producto.clasificacion == '3-EMPAQUE'):
#                         producto.sku_new = '3-'+str(producto.sku)
#                         producto.tipo = 'variation'
#                         producto.superior_new = producto.superior
#                         producto.posicion = contador_posicion
#                         contador_posicion += 1
#                 else:
#                     producto.superior_new = producto.superior
#                     producto.tipo = 'variation'
#                     producto.sku_new = str(producto.sku)        # Sin cambios en SKU
#                     producto.posicion = contador_posicion
#                     contador_posicion += 1

        #print_productos(familia_actual)
        #  
# print('TIPOS NUEVOS')
# for producto in lista_productos:
#     print(producto.tipo)

# print('SKU NUEVOS')
# for producto in lista_productos:
#     print(producto.sku_new)

# print('POSICIONES')
# for producto in lista_productos:
#    print(producto.posicion)

#print('SUPERIORES NUEVOS')
#for producto in lista_productos:
    #print(producto.superior_new)

# for producto in lista_productos:
#     print(str(producto.id) +' '+ str(producto.posicion))

#lista_productos_padres = get_productos_padres(lista_productos)
#print(len(lista_productos_padres))
#print_productos(lista_productos_padres)



################################## Obtención de presentaciones (valor de atributo 2) ################################################################################
# for producto in lista_productos:
#     lista_presentaciones = []
#     res = []
#     if(producto.tipo == 'variable' or producto.tipo == 'simple'):
#         familia_actual = get_familia(producto.sku)
#         for prod in familia_actual:
#             #print(prod.unidad,end=', ')
#             lista_presentaciones.append(prod.unidad)
#         [res.append(x) for x in lista_presentaciones if x not in res]
#         for item in res:
#             if(item == res[-1]):    # Último elemento de la lista
#                 print(item)
#             else:
#                 print(item,end=', ')
#     else:
#         print(producto.atributo)

#Obtención de atributos concatenados para valor atributo 3
# for producto in lista_productos:
#     lista_presentaciones = []
#     res = []
#     if(producto.tipo == 'variable' or producto.tipo == 'simple'):
#         familia_actual = get_familia(producto.sku)
#         for prod in familia_actual:
#             print(prod.atributo,end=', ')
#     else:
#         pass
#     print('')

    

#Obtención de tamaño de familia
# lista_sku_familias.sort()
# # for sku in lista_sku_familias:
# #     print(sku)
# for producto in lista_productos:
#     familia_prod = get_familia(producto.superior)
#     if(len(familia_prod)<=4):
#         print(producto.superior)
#     else:
#         print('grande')

except Exception as ex:
    print(str(ex))

##################################################################### Inicio de programa ###################################################################################
# if __name__ == "__main__":
#     main()